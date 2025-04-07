
from openpyxl import Workbook, load_workbook
import os
from datetime import date, datetime
import shutil

# Keep track of the current session
CURRENT_SESSION_ID = datetime.now().strftime("%Y%m%d%H%M%S")
ATTENDANCE_FILE = "attendance.xlsx"
ARCHIVE_DIR = "attendance_archives"

def initialize_attendance_file():
    """Create a new attendance file with headers"""
    wb = Workbook()
    sheet = wb.active
    sheet.title = "Attendance"
    sheet.append(["Date", "Session", "Name", "Status", "Engagement", "Remarks", "Posture"])
    
    # Ensure we have the archive directory
    if not os.path.exists(ARCHIVE_DIR):
        os.makedirs(ARCHIVE_DIR)
        
    # Save the new file
    wb.save(ATTENDANCE_FILE)
    print(f"Created new attendance file for session {CURRENT_SESSION_ID}")
    return wb

def update_attendance(name, engagement, remarks, posture_status):
    """Update the attendance Excel file with student data"""
    try:
        today = date.today().strftime("%Y-%m-%d")
        
        # Create file if it doesn't exist
        if not os.path.exists(ATTENDANCE_FILE):
            wb = initialize_attendance_file()
        else:
            try:
                wb = load_workbook(ATTENDANCE_FILE)
            except Exception as e:
                print(f"Error loading existing attendance file: {e}")
                # Backup the corrupted file if it exists
                if os.path.exists(ATTENDANCE_FILE):
                    corrupted_file = f"{ARCHIVE_DIR}/corrupted_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
                    try:
                        shutil.copy(ATTENDANCE_FILE, corrupted_file)
                        print(f"Backed up corrupted file to {corrupted_file}")
                    except Exception as backup_error:
                        print(f"Failed to backup corrupted file: {backup_error}")
                # Create a new file
                wb = initialize_attendance_file()
        
        sheet = wb.active
        
        # Append new data including session ID
        sheet.append([today, CURRENT_SESSION_ID, name, "Present", engagement, remarks, posture_status])
        wb.save(ATTENDANCE_FILE)
        print(f"Updated attendance for {name} in session {CURRENT_SESSION_ID}")

    except Exception as e:
        print(f"Error updating attendance: {e}")

def get_attendance_records(session_id=None):
    """Get attendance records from the Excel file, optionally filtered by session"""
    if not os.path.exists(ATTENDANCE_FILE):
        print("No attendance file found, creating a new one")
        initialize_attendance_file()
        return []

    try:
        wb = load_workbook(ATTENDANCE_FILE)
        sheet = wb.active

        data = []
        for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header
            if len(row) < 7:  # Skip incomplete rows
                continue
                
            # Check if we need to filter by session
            if session_id and row[1] != session_id:
                continue
                
            data.append({
                'date': row[0],
                'session': row[1],
                'name': row[2],
                'status': row[3],
                'engagement': row[4],
                'remarks': row[5],
                'posture': row[6]
            })

        return data
    except Exception as e:
        print(f"Error in get_attendance_records: {e}")
        return []

def reset_session():
    """Start a new session by updating the session ID"""
    global CURRENT_SESSION_ID
    
    # Archive the current session file if it exists
    if os.path.exists(ATTENDANCE_FILE):
        try:
            # Create the archive directory if it doesn't exist
            if not os.path.exists(ARCHIVE_DIR):
                os.makedirs(ARCHIVE_DIR)
                
            archive_path = f"{ARCHIVE_DIR}/attendance_{CURRENT_SESSION_ID}.xlsx"
            shutil.copy(ATTENDANCE_FILE, archive_path)
            print(f"Archived attendance for session {CURRENT_SESSION_ID} to {archive_path}")
        except Exception as e:
            print(f"Error archiving attendance: {e}")
    
    # Update the session ID
    CURRENT_SESSION_ID = datetime.now().strftime("%Y%m%d%H%M%S")
    print(f"Starting new session: {CURRENT_SESSION_ID}")
    
    # Create a fresh attendance file
    initialize_attendance_file()
    
    return CURRENT_SESSION_ID

def get_current_session_id():
    """Return the current session ID"""
    return CURRENT_SESSION_ID

